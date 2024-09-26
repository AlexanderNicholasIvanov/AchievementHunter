from bs4 import BeautifulSoup
import requests, html5lib
from openpyxl import Workbook, load_workbook

class Achievements:
    def __init__(self, steam_id, game_name):
        self.steam_id = steam_id
        self.game_name = game_name

    def get_page(self, game_id):
        url = f"https://steamcommunity.com/id/{self.steam_id}/stats/{game_id}/achievements/"
        r = requests.get(url)
        soup = BeautifulSoup(r.content, 'html.parser')
        return soup

    def get_game_id(self):
        url = "https://api.steampowered.com/ISteamApps/GetAppList/v2/"
        response = requests.get(url)
        app_list = response.json()
        app_list = sorted(app_list['applist']['apps'], key=lambda x: x['appid'])

        relevent_list = {}
        counter = 1
        for app in app_list:
            if self.game_name.lower() in app['name'].lower():
                relevent_list[app['appid']] = app['name']
        for i in relevent_list:
            print(str(counter)+") "+str(i)+" :: "+relevent_list[i])
            counter+=1

        option = int(input("Select an option: "))
        while option-1 > len(relevent_list) or option-1 < 0:
            option = int(input("Select an option: "))
        return list(relevent_list.keys())[option-1]

    def get_achievements(self, soup):
        category = soup.findAll("div", class_ = "achieveRow")
        return category

    def get_title(self, soup):
        header = soup.find("title")
        print(header)
        return header.text.strip().split(" :: ")[1]

    def get_achievement_and_progress(self, achievements):
        firstLast = []
        achievements = achievements[::-1]
        for achievement in achievements:
            achievement = [" ".join(i.split("\t")).strip() for i in achievement.text.strip().split("\n") if i and i != "\r"]
            firstLast.append(achievement)
        return firstLast

    def write_to_excel(self, title, achievements_array, ach_all, ach_comp):
        file = "C:\\Users\\Alex\\Documents\\Achievements.xlsx"
        try:
            workbook = load_workbook(file)
            if title in workbook.sheetnames:
                sheet = workbook[title]
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sheet = workbook.create_sheet(title)
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active()
            sheet.title = title

        sheet.cell(row=1, column=1).value = title
        sheet.cell(row=2, column=1).value = "Progress: " + str(ach_comp) + "/" + str(ach_all)
        sheet.cell(row=2, column=2).value = "{:.2f}".format(ach_comp / ach_all * 100) + "%" if ach_all != 0 else "0%"
        row_index = 4

        for row in achievements_array:
            sheet.cell(row=row_index, column=1).value = "✅" if "Unlocked" in row[len(row)-1] else "❌"
            for i in range(len(row)):
                sheet.cell(row=row_index, column=2+i).value = row[i]
            row_index+=1

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B', 'C')

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Add some padding to the width
            adjusted_width = max_length
            sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(file)

def main():
    steam_id = "theonetrueorphan"
    game_name = input("Enter game name: ")

    achievements    = Achievements(steam_id, game_name)
    game_id         = achievements.get_game_id()
    soup            = achievements.get_page(game_id)
    game_title      = achievements.get_title(soup)
    all_achieve     = achievements.get_achievements(soup)
    first_last      = achievements.get_achievement_and_progress(all_achieve)

    all_achievements_val    = len(first_last)
    achievements_completed  = 0
    for achievement in first_last:
        if "Unlocked" in achievement[len(achievement)-1]:
            achievements_completed +=1

    achievements.write_to_excel(game_title, first_last, all_achievements_val, achievements_completed)

main()