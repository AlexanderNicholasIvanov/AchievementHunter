from bs4 import BeautifulSoup
import requests, html5lib
from openpyxl import Workbook, load_workbook
import os

STEAM_PROFILE_ID_MIN_LENGTH = 17 # if your profile number is shorter than 17 characters, then you'll have to edit this number
class Achievements:
    def __init__(self, steam_id, game_name):
        self.steam_id = steam_id
        self.game_name = game_name

    def get_page(self, game_id):

        # is this a custome id or a profile id
        if self.steam_id.isdigit() and len(self.steam_id) >= STEAM_PROFILE_ID_MIN_LENGTH:
            url = f"https://steamcommunity.com/profiles/{self.steam_id}/stats/{game_id}/?tab=achievements"
        else:
            url = f"https://steamcommunity.com/id/{self.steam_id}/stats/{game_id}/achievements"
        r = requests.get(url)
        soup = BeautifulSoup(r.content, 'html.parser')
        return soup

    # function retrieves the game id from the name of the game (and user's choice)
    def get_game_id_and_title(self):
        url = "https://api.steampowered.com/ISteamApps/GetAppList/v2/"
        response = requests.get(url)
        app_list = response.json()
        app_list = sorted(app_list['applist']['apps'], key=lambda x: x['appid'])
        relevent_list = {}
        counter = 1

        # iterate through the json game data to retrieve game names that are relevant to what the user searched
        for app in app_list:
            if self.game_name.lower() in app['name'].lower():
                relevent_list[app['appid']] = app['name']

        # display the options that are relevant to the search
        for i in relevent_list:
            print(str(counter)+") "+str(i)+" :: "+relevent_list[i])
            counter+=1

        # user is prompted to select a valid option from the list
        option = int(input("Select an option: "))
        while option-1 > len(relevent_list) or option-1 < 0:
            option = int(input("Select an option: "))
        game_id = list(relevent_list.keys())[option-1]

        # return game id and title
        return game_id, "".join(relevent_list[game_id].split(":"))

    # this will get all the unfiltered achievements (will be cleaned up later)
    def get_achievements(self, soup):
        category = soup.findAll("div", class_ = "achieveRow")
        return category

    # this returns a double array with data containing Name of the achievement, description (if it has one) and date unlocked
    def get_achievement_and_progress(self, achievements):
        firstLast = []
        achievements = achievements[::-1]
        for achievement in achievements:
            # remove any garbage data or strings or any special characters (in this case \n, \r and \t)
            achievement = [" ".join(i.split("\t")).strip() for i in achievement.text.strip().split("\n") if i and i != "\r"]
            firstLast.append(achievement)
        return firstLast

    # Data is written to Achievements.xlsx in your Documents folder in your C drive (change the location to where ever you like)
    def write_to_excel(self, title, achievements_array, ach_all, ach_comp):
        file = f"C:\\Users\\{os.getlogin()}\\Documents\\Achievements-{self.steam_id}.xlsx"
        try:
            workbook = load_workbook(file)
            if title in workbook.sheetnames:
                sheet = workbook[title]

                # before writing, make sure to clear sheet. Wouldn't want old data mixed in with new data right?
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sheet = workbook.create_sheet(title)
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = title

        # using some of the top row cells to fill in the title of the game, progress, and progress completion as a percentage
        sheet.cell(row=1, column=1).value = title
        sheet.cell(row=2, column=1).value = "Progress: " + str(ach_comp) + "/" + str(ach_all)
        sheet.cell(row=2, column=2).value = "{:.2f}".format(ach_comp / ach_all * 100) + "%" if ach_all != 0 else "0%"
        row_index = 4

        # filling in rows and columns with appropriate data
        for row in achievements_array:
            sheet.cell(row=row_index, column=1).value = "✅" if "Unlocked" in row[len(row)-1] else "❌"
            for i in range(len(row)):
                sheet.cell(row=row_index, column=2+i).value = row[i]
            row_index+=1

        # adjusting the size of the cell blocks relative to the largest block in the column
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B', 'C')

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Add some padding to the width
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(file)

def main():
    steam_id = input("Enter you steam id: ")
    game_name = input("Enter game name: ")

    achievements        = Achievements(steam_id, game_name)
    game_id, game_title = achievements.get_game_id_and_title()
    soup                = achievements.get_page(game_id)
    all_achieve         = achievements.get_achievements(soup)
    first_last          = achievements.get_achievement_and_progress(all_achieve)

    all_achievements_val    = len(first_last)
    achievements_completed  = 0
    for achievement in first_last:
        if "Unlocked" in achievement[len(achievement)-1]:
            achievements_completed +=1

    achievements.write_to_excel(game_title, first_last, all_achievements_val, achievements_completed)

main()
